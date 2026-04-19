"""
Training & Placement Officer Router.
Handles all T&P related operations for companies, drives, and student applications.
"""

from fastapi import APIRouter, Depends, UploadFile, File, Form, HTTPException
from sqlalchemy.ext.asyncio import AsyncSession
from typing import List
import openpyxl
import io

from database import get_db
from app.core.security import require_role
from app.core.response import mark_enveloped
from app.services.tpo_service import TPOService

router = APIRouter()

def get_tpo_service(session: AsyncSession = Depends(get_db)):
    return TPOService(session)


@router.get("/tpo/companies", dependencies=[Depends(mark_enveloped)])
async def get_companies(
    user: dict = Depends(require_role("tpo", "tp_officer", "admin")),
    svc: TPOService = Depends(get_tpo_service)
):
    return {"data": await svc.get_companies(user["college_id"])}


@router.post("/tpo/companies")
async def create_company(
    data: dict,
    user: dict = Depends(require_role("tpo", "tp_officer", "admin")),
    svc: TPOService = Depends(get_tpo_service)
):
    uid = await svc.create_company(user["college_id"], data)
    return {"success": True, "id": uid}


@router.get("/tpo/drives", dependencies=[Depends(mark_enveloped)])
async def get_drives(
    user: dict = Depends(require_role("tpo", "tp_officer", "admin", "student")),
    svc: TPOService = Depends(get_tpo_service)
):
    return {"data": await svc.get_drives(user["college_id"])}


@router.post("/tpo/drives")
async def create_drive(
    data: dict,
    user: dict = Depends(require_role("tpo", "tp_officer", "admin")),
    svc: TPOService = Depends(get_tpo_service)
):
    uid = await svc.create_drive(user["college_id"], data)
    return {"success": True, "id": uid}


@router.put("/tpo/drives/{drive_id}")
async def update_drive(
    drive_id: str,
    data: dict,
    user: dict = Depends(require_role("tpo", "tp_officer", "admin")),
    svc: TPOService = Depends(get_tpo_service)
):
    await svc.update_drive(user["college_id"], drive_id, data)
    return {"success": True}


@router.get("/tpo/drives/{drive_id}/applicants", dependencies=[Depends(mark_enveloped)])
async def get_applicants(
    drive_id: str,
    user: dict = Depends(require_role("tpo", "tp_officer", "admin")),
    svc: TPOService = Depends(get_tpo_service)
):
    return {"data": await svc.get_applicants(user["college_id"], drive_id)}


@router.put("/tpo/drives/{drive_id}/shortlist")
async def shortlist_bulk(
    drive_id: str,
    payload: dict,
    user: dict = Depends(require_role("tpo", "tp_officer", "admin")),
    svc: TPOService = Depends(get_tpo_service)
):
    # payload expects { "student_ids": [...] }
    count = await svc.shortlist_bulk(user["college_id"], drive_id, payload.get("student_ids", []))
    return {"success": True, "shortlisted": count}


@router.put("/tpo/drives/{drive_id}/results")
async def log_result(
    drive_id: str,
    data: dict,
    user: dict = Depends(require_role("tpo", "tp_officer", "admin")),
    svc: TPOService = Depends(get_tpo_service)
):
    await svc.log_result(user["college_id"], drive_id, data)
    return {"success": True}


@router.put("/tpo/drives/{drive_id}/select")
async def select_candidate(
    drive_id: str,
    data: dict,
    user: dict = Depends(require_role("tpo", "tp_officer", "admin")),
    svc: TPOService = Depends(get_tpo_service)
):
    await svc.select_candidate(user["college_id"], drive_id, data)
    return {"success": True}


@router.get("/tpo/statistics")
async def get_stats(
    user: dict = Depends(require_role("tpo", "tp_officer", "admin", "principal")),
    svc: TPOService = Depends(get_tpo_service)
):
    return await svc.get_stats(user["college_id"])


# ════════════════════════════════════════════════════════════════════════════════
# Placement Restrictions (Blacklist)
# ════════════════════════════════════════════════════════════════════════════════

@router.get("/tpo/restrictions", dependencies=[Depends(mark_enveloped)])
async def get_restrictions(
    user: dict = Depends(require_role("tpo", "tp_officer", "admin")),
    svc: TPOService = Depends(get_tpo_service)
):
    return {"data": await svc.get_restrictions(user["college_id"])}


@router.get("/tpo/restrictions/student/{student_id}", dependencies=[Depends(mark_enveloped)])
async def get_student_restrictions(
    student_id: str,
    user: dict = Depends(require_role("tpo", "tp_officer", "admin")),
    svc: TPOService = Depends(get_tpo_service)
):
    return {"data": await svc.get_student_restrictions(user["college_id"], student_id)}


@router.post("/tpo/restrictions")
async def add_restriction(
    data: dict,
    user: dict = Depends(require_role("tpo", "tp_officer", "admin")),
    svc: TPOService = Depends(get_tpo_service)
):
    """
    Expects data: { student_id, reason, restriction_type, drive_id (optional), expires_at (optional) }
    """
    res_id = await svc.add_restriction(user["college_id"], user["id"], data)
    return {"success": True, "id": res_id}


@router.delete("/tpo/restrictions/{restriction_id}")
async def remove_restriction(
    restriction_id: str,
    user: dict = Depends(require_role("tpo", "tp_officer", "admin")),
    svc: TPOService = Depends(get_tpo_service)
):
    await svc.remove_restriction(user["college_id"], restriction_id)
    return {"success": True}


# ════════════════════════════════════════════════════════════════════════════════
# Excel Upload — Upload eligible students for a drive
# ════════════════════════════════════════════════════════════════════════════════

@router.post("/tpo/drives/{drive_id}/upload-students")
async def upload_students(
    drive_id: str,
    file: UploadFile = File(...),
    roll_column: int = Form(0),  # 0-indexed column number containing roll numbers
    user: dict = Depends(require_role("tpo", "tp_officer", "admin")),
    svc: TPOService = Depends(get_tpo_service)
):
    """
    Upload an Excel file with eligible/shortlisted students.
    The `roll_column` parameter specifies which column (0-indexed) contains roll numbers.
    """
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Only .xlsx/.xls files are supported")

    contents = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(contents), read_only=True, data_only=True)
        ws = wb.active
    except Exception:
        raise HTTPException(status_code=400, detail="Could not parse the Excel file")

    # Extract roll numbers from the specified column
    roll_numbers = []
    for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
        if row_idx == 0:
            continue  # Skip header row
        if roll_column < len(row) and row[roll_column]:
            val = str(row[roll_column]).strip()
            if val:
                roll_numbers.append(val)
    wb.close()

    if not roll_numbers:
        raise HTTPException(status_code=400, detail="No roll numbers found in the specified column")

    result = await svc.upload_eligible_students(user["college_id"], drive_id, roll_numbers)
    return {"success": True, **result}


@router.post("/tpo/drives/{drive_id}/preview-excel")
async def preview_excel(
    drive_id: str,
    file: UploadFile = File(...),
    user: dict = Depends(require_role("tpo", "tp_officer", "admin")),
):
    """
    Preview Excel file headers and first 5 rows to help TPO pick the right column.
    """
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Only .xlsx/.xls files are supported")

    contents = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(contents), read_only=True, data_only=True)
        ws = wb.active
    except Exception:
        raise HTTPException(status_code=400, detail="Could not parse the Excel file")

    headers = []
    preview_rows = []
    total_rows = 0

    for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
        if row_idx == 0:
            headers = [str(c) if c else f"Column {i+1}" for i, c in enumerate(row)]
        elif row_idx <= 5:
            preview_rows.append([str(c) if c else "" for c in row])
        total_rows = row_idx

    wb.close()

    return {
        "headers": headers,
        "preview_rows": preview_rows,
        "total_data_rows": total_rows,  # excluding header
    }
